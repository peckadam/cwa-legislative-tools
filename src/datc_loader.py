"""
datc_loader.py
--------------
Loads and structures the CWA Day at the Capitol (DATC) Meeting Matrix
from the Excel workbook for display in the Streamlit app.

Primary output: a matrix DataFrame (legislators × attending WDAs) with
cell values of "Own", "Neighbor", or "" indicating meeting type.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd


# ---------------------------------------------------------------------------
# WDA display configuration
# ---------------------------------------------------------------------------

# Canonical short names as they appear in the "Legislator Coverage" sheet,
# in the presentation order derived from the Master Summary sheet.
WDA_DISPLAY_ORDER: List[str] = [
    "SELACO WDB",
    "San Luis Obispo WDB",
    "San Bernardino County WDB",
    "North Central Counties",
    "San Francisco OEWD",
    "South Bay WIB",
    "Tulare County WIB",
    "Ventura County WDB",
    "Anaheim WC",
    "Richmond E&T",
    "Fresno Regional WDB",
    "San Joaquin County WorkNet",
    "Humboldt County",
    "Long Beach WIN",
    "Alameda County WDB",
    "Mother Lode",
    "Imperial County",
    "Solano County WDB",
    "North Bay Alliance",
    "San Diego WP",
]

# Abbreviated column labels (≤10 chars) for wide-table display.
WDA_ABBREVIATIONS: Dict[str, str] = {
    "SELACO WDB": "SELACO",
    "San Luis Obispo WDB": "SLO",
    "San Bernardino County WDB": "San Bern.",
    "North Central Counties": "N. Central",
    "San Francisco OEWD": "SF OEWD",
    "South Bay WIB": "South Bay",
    "Tulare County WIB": "Tulare",
    "Ventura County WDB": "Ventura",
    "Anaheim WC": "Anaheim",
    "Richmond E&T": "Richmond",
    "Fresno Regional WDB": "Fresno",
    "San Joaquin County WorkNet": "San Joaquin",
    "Humboldt County": "Humboldt",
    "Long Beach WIN": "Long Beach",
    "Alameda County WDB": "Alameda",
    "Mother Lode": "Mother Lode",
    "Imperial County": "Imperial",
    "Solano County WDB": "Solano",
    "North Bay Alliance": "N. Bay",
    "San Diego WP": "San Diego",
}

# Region grouping for colour-coded column headers.
WDA_REGIONS: Dict[str, str] = {
    "SELACO WDB": "Southern CA",
    "San Luis Obispo WDB": "Central Coast",
    "San Bernardino County WDB": "Inland Empire",
    "North Central Counties": "Northern CA",
    "San Francisco OEWD": "Bay Area",
    "South Bay WIB": "Southern CA",
    "Tulare County WIB": "Central Valley",
    "Ventura County WDB": "Central Coast",
    "Anaheim WC": "Orange County",
    "Richmond E&T": "Bay Area",
    "Fresno Regional WDB": "Central Valley",
    "San Joaquin County WorkNet": "Central Valley",
    "Humboldt County": "Northern CA",
    "Long Beach WIN": "Southern CA",
    "Alameda County WDB": "Bay Area",
    "Mother Lode": "Northern CA",
    "Imperial County": "Border",
    "Solano County WDB": "Bay Area",
    "North Bay Alliance": "Bay Area",
    "San Diego WP": "San Diego",
}

# Cell display symbols
OWN_SYMBOL = "✓"
NEIGHBOR_SYMBOL = "~"
NO_COVERAGE_LABEL = "No coverage"


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def _parse_legislator_coverage(ws) -> List[Dict[str, Any]]:
    """Parse the 'Legislator Coverage' worksheet into a list of dicts."""
    legislators: List[Dict[str, Any]] = []
    chamber: Optional[str] = None

    for row in ws.iter_rows(values_only=True):
        val0 = row[0]
        if val0 == "ASSEMBLY MEMBERS":
            chamber = "Assembly"
            continue
        if val0 == "SENATORS":
            chamber = "Senate"
            continue
        # Skip header rows and blanks
        if val0 in (None, "District", "COVERAGE STATISTICS") or not chamber:
            continue
        # Skip summary lines (e.g. "Assembly Districts with meetings: ...")
        if isinstance(val0, str) and not val0.startswith(("AD-", "SD-")):
            continue

        district = str(val0).strip()
        member = str(row[1]).strip() if row[1] else ""
        party = str(row[2]).strip() if row[2] else ""
        visited_by_raw = str(row[3]).strip() if row[3] else ""
        meeting_types_raw = str(row[4]).strip() if row[4] else ""
        priority = str(row[5]).strip() if row[5] else "General"
        must_schedule = str(row[6]).strip() if row[6] else "No"

        if visited_by_raw in ("", "NO COVERAGE", "None"):
            visited_by_list: List[str] = []
            meeting_types_list: List[str] = []
            has_coverage = False
        else:
            visited_by_list = [x.strip() for x in visited_by_raw.split(";") if x.strip()]
            meeting_types_list = [x.strip() for x in meeting_types_raw.split(";") if x.strip()]
            has_coverage = True

        # Build per-WDA meeting type map for this legislator
        wda_meetings: Dict[str, str] = {}
        for i, wda in enumerate(visited_by_list):
            mtype = meeting_types_list[i] if i < len(meeting_types_list) else "Unknown"
            wda_meetings[wda] = mtype

        legislators.append(
            {
                "chamber": chamber,
                "district": district,
                "member": member,
                "party": party,
                "priority": priority if priority and priority != "None" else "General",
                "must_schedule": must_schedule if must_schedule and must_schedule != "None" else "No",
                "has_coverage": has_coverage,
                "wda_meetings": wda_meetings,
            }
        )

    return legislators


def _parse_master_summary(ws) -> Dict[str, Dict[str, Any]]:
    """
    Parse the 'Master Summary' sheet.
    Returns a dict keyed by WDA long name → {attendees, region}.
    """
    result: Dict[str, Dict[str, Any]] = {}
    in_data = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == "Workforce Area":
            in_data = True
            continue
        if not in_data:
            continue
        if row[0] in (None, "TOTALS"):
            continue
        wda_long = str(row[0]).strip()
        attendees = row[1] if row[1] is not None else 0
        region = str(row[7]).strip() if row[7] else ""
        result[wda_long] = {"attendees": attendees, "region": region}
    return result


def _parse_area_attendees(wb, area_sheets: List[str]) -> Dict[str, List[str]]:
    """
    Parse attendee names from individual WDA sheets.
    Returns {wda_short_name: [attendee strings]}.
    """
    attendees: Dict[str, List[str]] = {}
    for sheet_name in area_sheets:
        ws = wb[sheet_name]
        people: List[str] = []
        in_attendees = False
        for row in ws.iter_rows(values_only=True):
            if row[0] == "ATTENDEES":
                in_attendees = True
                continue
            if not in_attendees:
                continue
            if row[0] is None:
                continue
            name_part = str(row[0]).strip()
            # Stop when we hit summary lines or meeting assignment tables
            if (
                name_part in ("Legislator", "District", "ASSEMBLY MEETINGS", "SENATE MEETINGS")
                or name_part.startswith(("AD-", "SD-", "TOTAL MEETINGS", "TOTAL:"))
            ):
                break
            title_part = str(row[1]).strip() if row[1] else ""
            if name_part:
                person = name_part + (f", {title_part}" if title_part else "")
                people.append(person)
        attendees[sheet_name] = people
    return attendees


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def load_datc_data(excel_path: Path) -> Dict[str, Any]:
    """
    Load and structure all DATC data from the Meeting Matrix workbook.

    Returns a dict with:
      - "legislators"   : list of dicts with per-legislator coverage info
      - "matrix"        : wide DataFrame (legislators × WDA columns)
      - "wda_order"     : ordered list of WDA short names (canonical)
      - "wda_abbrevs"   : {wda_name: abbreviation} for column headers
      - "wda_regions"   : {wda_name: region}
      - "area_attendees": {wda_name: [attendee strings]}
      - "summary_stats" : dict with totals for banner
      - "data_loaded"   : bool (False if file missing)
    """
    if not excel_path.exists():
        return {"data_loaded": False}

    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception:
        return {"data_loaded": False}

    # ---- Parse core sheets ------------------------------------------------
    legislators = _parse_legislator_coverage(wb["Legislator Coverage"])

    area_sheets = [s for s in wb.sheetnames if not s.startswith("CUT - ")
                   and s not in ("Master Summary", "Legislator Coverage")]

    area_attendees_raw = _parse_area_attendees(wb, area_sheets)
    # Re-key by the WDA display name (strip sheet name prefixes)
    # Sheet names for individual areas match the short names directly.
    area_attendees: Dict[str, List[str]] = {}
    for sheet_name, people in area_attendees_raw.items():
        # Sheet titles like "Day at the Capitol 2026 — Contra Costa County WDB"
        # are in row[0] of the sheet; use sheet_name directly as key fallback.
        # Map to WDA_DISPLAY_ORDER names via fuzzy match on sheet name
        matched = _match_area_name(sheet_name)
        if matched:
            area_attendees[matched] = people

    # ---- Build matrix DataFrame --------------------------------------------
    # Determine which WDAs actually appear (preserve display order)
    all_wdas_in_data: set = set()
    for leg in legislators:
        all_wdas_in_data.update(leg["wda_meetings"].keys())

    # Use canonical order, filtered to what's present
    wda_order = [w for w in WDA_DISPLAY_ORDER if w in all_wdas_in_data]
    # Append any unexpected names at the end
    for w in sorted(all_wdas_in_data):
        if w not in wda_order:
            wda_order.append(w)

    # Build rows
    matrix_rows = []
    for leg in legislators:
        row: Dict[str, Any] = {
            "Chamber": leg["chamber"],
            "District": leg["district"],
            "Legislator": leg["member"],
            "Party": leg["party"],
            "Priority": leg["priority"],
            "Must Schedule": leg["must_schedule"],
        }
        for wda in wda_order:
            mtype = leg["wda_meetings"].get(wda, "")
            if mtype == "Own":
                row[wda] = OWN_SYMBOL
            elif mtype == "Neighbor":
                row[wda] = NEIGHBOR_SYMBOL
            else:
                row[wda] = ""
        matrix_rows.append(row)

    matrix = pd.DataFrame(matrix_rows)

    # ---- Summary stats -----------------------------------------------------
    covered = sum(1 for leg in legislators if leg["has_coverage"])
    must_sched = sum(1 for leg in legislators if leg["must_schedule"] == "Yes")
    high_priority = sum(1 for leg in legislators if leg["priority"] == "High")
    total_attending = sum(
        len(v) for v in area_attendees.values()
    )

    summary_stats = {
        "total_areas": len(wda_order),
        "total_legislators": len(legislators),
        "covered_legislators": covered,
        "must_schedule": must_sched,
        "high_priority": high_priority,
        "total_attendees": total_attending,
    }

    return {
        "data_loaded": True,
        "legislators": legislators,
        "matrix": matrix,
        "wda_order": wda_order,
        "wda_abbrevs": WDA_ABBREVIATIONS,
        "wda_regions": WDA_REGIONS,
        "area_attendees": area_attendees,
        "summary_stats": summary_stats,
    }


def _match_area_name(sheet_name: str) -> Optional[str]:
    """Fuzzy-match an individual area sheet name to a canonical WDA name."""
    # Direct match first
    if sheet_name in WDA_DISPLAY_ORDER:
        return sheet_name
    # Substring match
    for canonical in WDA_DISPLAY_ORDER:
        if canonical.lower() in sheet_name.lower() or sheet_name.lower() in canonical.lower():
            return canonical
    # Special cases
    special: Dict[str, str] = {
        "Contra Costa County WDB": "Alameda County WDB",  # not in DATC list; skip
        "San Joaquin County WorkNet": "San Joaquin County WorkNet",
    }
    return special.get(sheet_name)


# ---------------------------------------------------------------------------
# Styling helper
# ---------------------------------------------------------------------------

def style_matrix(df: pd.DataFrame, wda_order: List[str]) -> "pd.io.formats.style.Styler":
    """Apply background colours to the matrix for display."""
    own_bg = "#c8e6c9"       # light green
    neighbor_bg = "#fff9c4"  # light yellow
    none_bg = "#f5f5f5"      # light grey for no-coverage rows

    def cell_style(val: str) -> str:
        if val == OWN_SYMBOL:
            return f"background-color: {own_bg}; text-align: center; font-weight: bold;"
        if val == NEIGHBOR_SYMBOL:
            return f"background-color: {neighbor_bg}; text-align: center;"
        return "text-align: center; color: #bbb;"

    # Identify which columns are WDA columns
    wda_cols = [c for c in df.columns if c in wda_order]
    styled = df.style.applymap(cell_style, subset=wda_cols)
    return styled
